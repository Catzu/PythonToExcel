# Installation
# pip install openpyxl
from openpyxl import Workbook, load_workbook
import os

# Creates excel file if it doesn't exist
if not os.path.exists("excel.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Data"
    wb.save("excel.xlsx")

# Load workbook and sheet
wk = load_workbook('excel.xlsx')
ws = wk.active

# Define header
header = ['Date', 'Sales Rep', 'Products', 'Units', 'Price']

# Check if header exists in the first row
header_exists = False
for row in ws.iter_rows(values_only=True):
    if row == tuple(header):
        header_exists = True
        break

# Check first row and write header directly into row 1 if needed
first_row = [cell.value for cell in ws[1]]
if first_row != header:
    for col, val in enumerate(header, start=1):
        ws.cell(row=1, column=col, value=val)

# Append data row (Data that is going into Excel)
ws.append(['05/04/2025', 'Kim Possible', 'Phone', 3, 600])
ws.append(['18/04/2025', 'Kraus Harvestein', 'Laptop', 9, 2900])

# Save workbook
wk.save('excel.xlsx')